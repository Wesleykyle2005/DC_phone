import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

def export_to_excel(datos, columnas, nombre_archivo="export.xlsx"):
    """
    Exporta una lista de diccionarios a un archivo Excel estilizado.
    - datos: lista de diccionarios (cada dict es una fila)
    - columnas: lista de tuplas ("nombre_columna", "título visible")
    - nombre_archivo: nombre sugerido para el archivo
    Retorna: BytesIO listo para ser enviado como respuesta de descarga.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell_alignment = Alignment(horizontal="center", vertical="center")

    # Escribir encabezados
    for col_idx, (col, titulo) in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = cell_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(titulo) + 2)

    # Escribir datos
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, (col, _) in enumerate(columnas, 1):
            valor = fila.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = cell_alignment

    # Ajustar altura de encabezado
    ws.row_dimensions[1].height = 25

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output 